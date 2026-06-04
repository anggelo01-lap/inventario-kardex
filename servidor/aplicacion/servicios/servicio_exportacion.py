from datetime import date, datetime
from io import BytesIO

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from aplicacion.servicios.servicio_movimiento import list_movimientos_filtrados, movimiento_a_lista_out
from aplicacion.servicios.servicio_producto import get_producto, list_productos


def _excel_safe_datetime(value: datetime | None):
    if value is None:
        return None
    if getattr(value, "tzinfo", None) is not None:
        return value.replace(tzinfo=None)
    return value


def _ws_apply_header_style(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="FFB91C1C")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="FFCBD5E1"),
        right=Side(style="thin", color="FFCBD5E1"),
        top=Side(style="thin", color="FFCBD5E1"),
        bottom=Side(style="thin", color="FFCBD5E1"),
    )

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border


def _ws_apply_body_style(ws, *, start_row: int, end_row: int, end_col: int) -> None:
    border = Border(
        left=Side(style="thin", color="FFE2E8F0"),
        right=Side(style="thin", color="FFE2E8F0"),
        top=Side(style="thin", color="FFE2E8F0"),
        bottom=Side(style="thin", color="FFE2E8F0"),
    )
    zebra = PatternFill("solid", fgColor="FFF8FAFC")
    body_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    body_font = Font(color="FF0F172A")

    for r in range(start_row, end_row + 1):
        row_fill = zebra if (r - start_row) % 2 == 1 else None
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = body_alignment
            cell.font = body_font
            if row_fill is not None:
                cell.fill = row_fill


def _ws_set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _ws_finalize_table(ws, *, header_row: int, last_row: int, last_col: int) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.row_dimensions[header_row].height = 22


def _tipo_color(tipo: str) -> PatternFill | None:
    t = (tipo or "").strip().lower()
    if t == "entrada":
        return PatternFill("solid", fgColor="FFDCFCE7")
    if t == "salida":
        return PatternFill("solid", fgColor="FFFEE2E2")
    if t == "ajuste":
        return PatternFill("solid", fgColor="FFEDE9FE")
    return None


def export_productos_xlsx(db: Session) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    headers = ["Codigo", "Nombre", "Descripcion", "Stock actual", "Stock minimo", "Precio (PEN)"]
    ws.append(headers)

    productos = list_productos(db)
    for p in productos:
        ws.append(
            [
                p.codigo,
                p.nombre,
                p.descripcion or "",
                p.stock_actual,
                p.stock_minimo,
                float(p.precio) if p.precio is not None else None,
            ]
        )

    last_row = ws.max_row
    last_col = len(headers)
    _ws_apply_header_style(ws, 1)
    _ws_set_widths(ws, [14, 36, 52, 14, 14, 14])
    _ws_finalize_table(ws, header_row=1, last_row=last_row, last_col=last_col)

    _ws_apply_body_style(ws, start_row=2, end_row=last_row, end_col=last_col)

    for r in range(2, last_row + 1):
        ws.cell(row=r, column=4).number_format = "0"
        ws.cell(row=r, column=5).number_format = "0"
        ws.cell(row=r, column=6).number_format = '"S/ " #,##0.00'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def export_movimientos_xlsx(
    db: Session,
    *,
    producto_id: int | None = None,
    tipo: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    limit: int = 5000,
) -> BytesIO:
    rows = list_movimientos_filtrados(
        db,
        producto_id=producto_id,
        tipo=tipo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        limit=limit,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    producto = get_producto(db, producto_id) if producto_id is not None else None
    producto_txt = f"{producto.codigo} - {producto.nombre}" if producto is not None else "Todos"
    tipo_txt = tipo if tipo else "Todos"
    desde_txt = fecha_desde.isoformat() if fecha_desde else "—"
    hasta_txt = fecha_hasta.isoformat() if fecha_hasta else "—"

    ws.append(["Movimientos (Kardex)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    accent = "FFB91C1C"
    ws["A1"].font = Font(bold=True, size=16, color=accent)
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFFFF")
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws["A1"].border = Border(bottom=Side(style="thick", color=accent))
    ws.row_dimensions[1].height = 28

    ws.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Producto", producto_txt, "Tipo", tipo_txt, "Desde", desde_txt, "Hasta", hasta_txt, "Registros", len(rows)])
    ws.append([])

    header_row = ws.max_row + 1
    headers = [
        "ID",
        "Fecha",
        "Codigo",
        "Producto",
        "Tipo",
        "Cantidad",
        "Cliente",
        "Proveedor",
        "Stock anterior",
        "Stock posterior",
        "Motivo",
        "Usuario",
    ]
    ws.append(headers)

    total_entradas = 0
    total_salidas = 0
    total_ajustes = 0
    for m in rows:
        dto = movimiento_a_lista_out(m)
        if dto.tipo == "entrada":
            total_entradas += 1
        elif dto.tipo == "salida":
            total_salidas += 1
        elif dto.tipo == "ajuste":
            total_ajustes += 1
        ws.append(
            [
                dto.id,
                _excel_safe_datetime(dto.fecha_movimiento),
                dto.producto_codigo or "",
                dto.producto_nombre or "",
                dto.tipo,
                dto.cantidad,
                dto.cliente_nombre or "",
                dto.proveedor_nombre or "",
                dto.stock_anterior,
                dto.stock_posterior,
                dto.motivo or "",
                dto.usuario_username,
            ]
        )

    ws.append([])
    ws.append(["Totales", "Entradas", total_entradas, "Salidas", total_salidas, "Ajustes", total_ajustes])

    last_row = ws.max_row - 2
    last_col = len(headers)
    _ws_apply_header_style(ws, header_row)
    _ws_set_widths(ws, [10, 20, 14, 40, 12, 10, 22, 22, 14, 14, 30, 18])
    _ws_finalize_table(ws, header_row=header_row, last_row=last_row, last_col=last_col)

    if last_row >= header_row + 1:
        _ws_apply_body_style(ws, start_row=header_row + 1, end_row=last_row, end_col=last_col)

    for r in range(header_row + 1, last_row + 1):
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=r, column=6).number_format = "0"
        ws.cell(row=r, column=9).number_format = "0"
        ws.cell(row=r, column=10).number_format = "0"
        tipo_cell = ws.cell(row=r, column=5)
        fill = _tipo_color(str(tipo_cell.value or ""))
        if fill is not None:
            tipo_cell.fill = fill
            tipo_cell.font = Font(bold=True, color="FF0F172A")
            tipo_cell.alignment = Alignment(vertical="center", horizontal="center")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _pdf_safe(text: str) -> str:
    return text.encode("latin-1", errors="replace").decode("latin-1")


def export_movimientos_pdf(
    db: Session,
    *,
    producto_id: int | None = None,
    tipo: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    limit: int = 500,
) -> BytesIO:
    rows = list_movimientos_filtrados(
        db,
        producto_id=producto_id,
        tipo=tipo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        limit=limit,
    )

    producto = get_producto(db, producto_id) if producto_id is not None else None
    producto_txt = f"{producto.codigo} - {producto.nombre}" if producto is not None else "Todos"
    tipo_txt = tipo if tipo else "Todos"
    desde_txt = fecha_desde.isoformat() if fecha_desde else "—"
    hasta_txt = fecha_hasta.isoformat() if fecha_hasta else "—"
    generado_txt = datetime.now().strftime("%Y-%m-%d %H:%M")

    total_entradas = 0
    total_salidas = 0
    total_ajustes = 0
    for m in rows:
        dto = movimiento_a_lista_out(m)
        if dto.tipo == "entrada":
            total_entradas += 1
        elif dto.tipo == "salida":
            total_salidas += 1
        elif dto.tipo == "ajuste":
            total_ajustes += 1

    class _MovimientosPdf(FPDF):
        def footer(self) -> None:
            self.set_y(-10)
            self.set_font("Helvetica", size=7)
            self.set_text_color(100, 116, 139)
            self.cell(0, 8, _pdf_safe(f"Página {self.page_no()}"), align="R")

    pdf = _MovimientosPdf(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(
        0,
        8,
        _pdf_safe("Movimientos de inventario (Kardex)"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.set_draw_color(185, 28, 28)
    pdf.set_line_width(0.6)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(2)

    pdf.set_font("Helvetica", size=8)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 5, _pdf_safe(f"Generado: {generado_txt}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 5, _pdf_safe(f"Producto: {producto_txt}  |  Tipo: {tipo_txt}  |  Desde: {desde_txt}  |  Hasta: {hasta_txt}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(
        0,
        5,
        _pdf_safe(
            f"Total: {len(rows)}  |  Entradas: {total_entradas}  |  Salidas: {total_salidas}  |  Ajustes: {total_ajustes}"
        ),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(2)

    pdf.set_font("Helvetica", size=7)
    col_w = [14, 28, 18, 52, 14, 12, 18, 30, 45, 24]
    headers = ["ID", "Fecha", "Cod.", "Producto", "Tipo", "Cant.", "Stock", "Tercero", "Motivo", "Usuario"]

    pdf.set_fill_color(241, 245, 249)
    pdf.set_text_color(15, 23, 42)
    pdf.set_draw_color(203, 213, 225)
    pdf.set_font("Helvetica", "B", 7)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 6, _pdf_safe(h), border=1, fill=True)
    pdf.ln()

    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", size=7)
    fill = False
    for m in rows:
        dto = movimiento_a_lista_out(m)
        tercero = dto.cliente_nombre or dto.proveedor_nombre or "-"
        stock_txt = (
            f"{dto.stock_anterior}->{dto.stock_posterior}"
            if dto.stock_anterior is not None and dto.stock_posterior is not None
            else "-"
        )
        motivo_txt = (dto.motivo or "—").replace("\n", " ").replace("\r", " ")
        if len(motivo_txt) > 70:
            motivo_txt = motivo_txt[:67] + "..."

        line = [
            str(dto.id),
            dto.fecha_movimiento.strftime("%Y-%m-%d %H:%M"),
            dto.producto_codigo or "",
            (dto.producto_nombre or "")[:45],
            dto.tipo,
            str(dto.cantidad),
            stock_txt,
            tercero[:28],
            motivo_txt,
            dto.usuario_username[:18],
        ]

        if fill:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)

        for i, cell in enumerate(line):
            if i == 4:
                if dto.tipo == "entrada":
                    pdf.set_text_color(22, 163, 74)
                elif dto.tipo == "salida":
                    pdf.set_text_color(220, 38, 38)
                elif dto.tipo == "ajuste":
                    pdf.set_text_color(124, 58, 237)
                else:
                    pdf.set_text_color(15, 23, 42)
            else:
                pdf.set_text_color(15, 23, 42)
            pdf.cell(col_w[i], 6, _pdf_safe(cell), border=1, fill=fill)
        pdf.ln()
        fill = not fill

    bio = BytesIO()
    pdf.output(bio)
    bio.seek(0)
    return bio
