import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from aplicacion.nucleo.base_datos import SessionLocal
from aplicacion.modelos.categoria import Categoria
from aplicacion.modelos.cliente import Cliente  # noqa: F401
from aplicacion.modelos.movimiento import Movimiento
from aplicacion.modelos.producto import Producto
from aplicacion.modelos.proveedor import Proveedor  # noqa: F401
from aplicacion.modelos.usuario import User  # noqa: F401


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Limpia productos/movimientos y carga productos desde un archivo Excel."
    )
    parser.add_argument(
        "--xlsx",
        default=str(Path(__file__).resolve().parents[2] / "database" / "DATA-PRODUCTOS.xlsx"),
        help="Ruta al archivo XLSX con columnas CODIGO y NOMBRE DEL PRODUCTO.",
    )
    parser.add_argument(
        "--categoria",
        default="General",
        help="Nombre de la categoria a asignar a los productos cargados.",
    )
    parser.add_argument(
        "--descripcion-categoria",
        default="Categoria generada para importacion desde Excel",
        help="Descripcion para crear la categoria si no existe.",
    )
    parser.add_argument(
        "--tipo",
        default="producto",
        choices=["producto", "repuesto", "insumo"],
        help="Tipo a asignar a todos los productos cargados.",
    )
    parser.add_argument("--stock-actual", type=int, default=0, help="Stock actual inicial.")
    parser.add_argument("--stock-minimo", type=int, default=0, help="Stock minimo inicial.")
    parser.add_argument("--precio", type=float, default=0.0, help="Precio inicial.")
    parser.add_argument(
        "--no-limpiar",
        action="store_true",
        help="No elimina datos existentes. Solo inserta/actualiza por codigo.",
    )
    return parser


def _normalize(value) -> str:
    return str(value).strip() if value is not None else ""


def read_rows(xlsx_path: Path) -> list[tuple[str, str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {xlsx_path}")

    workbook = load_workbook(filename=xlsx_path, data_only=True)
    worksheet = workbook.active

    header = [_normalize(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        codigo_idx = header.index("CODIGO")
        nombre_idx = header.index("NOMBRE DEL PRODUCTO")
    except ValueError as exc:
        raise ValueError(
            "No se encontraron las columnas esperadas 'CODIGO' y 'NOMBRE DEL PRODUCTO' en la primera fila."
        ) from exc

    rows: list[tuple[str, str]] = []
    seen_codes: set[str] = set()
    duplicated = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        codigo = _normalize(row[codigo_idx] if codigo_idx < len(row) else None)
        nombre = _normalize(row[nombre_idx] if nombre_idx < len(row) else None)
        if not codigo or not nombre:
            continue
        if codigo in seen_codes:
            duplicated += 1
            continue
        seen_codes.add(codigo)
        rows.append((codigo, nombre))

    if not rows:
        raise ValueError("El archivo no contiene filas validas para cargar.")

    if duplicated:
        print(f"Aviso: se omitieron {duplicated} codigos duplicados en el Excel.")
    return rows


def get_or_create_categoria(db, nombre: str, descripcion: str) -> Categoria:
    categoria = db.query(Categoria).filter(Categoria.nombre == nombre).first()
    if categoria is None:
        categoria = Categoria(nombre=nombre, descripcion=descripcion)
        db.add(categoria)
        db.flush()
    return categoria


def clear_product_data(db) -> tuple[int, int]:
    movimientos_eliminados = db.execute(delete(Movimiento)).rowcount or 0
    productos_eliminados = db.execute(delete(Producto)).rowcount or 0
    db.commit()
    return movimientos_eliminados, productos_eliminados


def upsert_productos(
    db,
    rows: list[tuple[str, str]],
    categoria_id: int,
    tipo: str,
    stock_actual: int,
    stock_minimo: int,
    precio: float,
) -> tuple[int, int]:
    creados = 0
    actualizados = 0
    for codigo, nombre in rows:
        producto = db.query(Producto).filter(Producto.codigo == codigo).first()
        if producto is None:
            producto = Producto(
                codigo=codigo,
                nombre=nombre,
                descripcion=None,
                categoria_id=categoria_id,
                proveedor_id=None,
                tipo=tipo,
                stock_actual=stock_actual,
                stock_minimo=stock_minimo,
                precio=precio,
            )
            db.add(producto)
            creados += 1
        else:
            producto.nombre = nombre
            producto.categoria_id = categoria_id
            producto.proveedor_id = None
            producto.tipo = tipo
            producto.stock_actual = stock_actual
            producto.stock_minimo = stock_minimo
            producto.precio = precio
            actualizados += 1
    db.commit()
    return creados, actualizados


def main() -> int:
    args = build_parser().parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()

    if args.stock_actual < 0 or args.stock_minimo < 0 or args.precio < 0:
        raise ValueError("stock_actual, stock_minimo y precio deben ser mayores o iguales a 0.")

    rows = read_rows(xlsx_path)
    db = SessionLocal()
    try:
        categoria = get_or_create_categoria(db, args.categoria.strip(), args.descripcion_categoria.strip())
        db.commit()

        movimientos_eliminados = 0
        productos_eliminados = 0
        if not args.no_limpiar:
            movimientos_eliminados, productos_eliminados = clear_product_data(db)
            # Categoria puede haber sido eliminada junto con productos? No, categorias se conserva.
            categoria = get_or_create_categoria(db, args.categoria.strip(), args.descripcion_categoria.strip())
            db.commit()

        creados, actualizados = upsert_productos(
            db=db,
            rows=rows,
            categoria_id=categoria.id,
            tipo=args.tipo,
            stock_actual=args.stock_actual,
            stock_minimo=args.stock_minimo,
            precio=args.precio,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    print(
        "Carga completada:\n"
        f"- Archivo: {xlsx_path}\n"
        f"- Filas validas: {len(rows)}\n"
        f"- Movimientos eliminados: {movimientos_eliminados}\n"
        f"- Productos eliminados: {productos_eliminados}\n"
        f"- Productos creados: {creados}\n"
        f"- Productos actualizados: {actualizados}\n"
        f"- Categoria asignada: {args.categoria.strip()}\n"
        f"- Tipo: {args.tipo}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
